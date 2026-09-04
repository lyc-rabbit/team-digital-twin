"""
日报增量同步服务

Excel 导入 = 数据同步任务
业务唯一键：report_date + member_id
Diff：NEW / UPDATED / UNCHANGED
变更触发 AI 标签分析；无变化跳过
"""

from __future__ import annotations

import hashlib
import io
import re
import threading
from collections import Counter
from datetime import datetime, date
from typing import Optional

from openpyxl import load_workbook

from database import (
    get_all_members,
    get_member,
    get_daily_reports_by_dates,
    get_daily_report_by_key,
    insert_daily_report,
    update_daily_report,
    create_daily_import_task,
    update_daily_import_task,
    get_daily_import_task,
    upsert_daily_report_analysis,
    get_member_recent_report_summary,
    get_daily_report_style_prompt,
    list_daily_report_style_prompts,
    update_daily_report_style_prompt,
)
from llm_client import analyze_daily_report, is_mock_mode, rewrite_daily_report, last_call_degraded
from timeutil import parse_day


# ========== Hash / 标准化 ==========

def normalize_content(content: str) -> str:
    """内容标准化：去空白、去标点、小写"""
    if content is None:
        return ""
    text = str(content).strip()
    text = re.sub(r"\s+", "", text)
    # 去掉常见中英文标点与符号（用 unicode 范围，避免转义问题）
    text = re.sub(r"[^\w\u4e00-\u9fff]+", "", text, flags=re.UNICODE)
    return text.lower()


def content_hash(content: str) -> str:
    return hashlib.md5(normalize_content(content).encode("utf-8")).hexdigest()


# ========== Excel 解析 ==========

def _cell_str(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def _parse_date(value) -> Optional[str]:
    raw = _cell_str(value)
    if not raw:
        return None
    # Excel 数字日期由 openpyxl 通常已转 datetime；兜底常见字符串
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(raw[:19], fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    # 仅日期前 10 位
    if re.match(r"^\d{4}-\d{2}-\d{2}", raw):
        return raw[:10]
    return None


def parse_daily_report_excel(file_bytes: bytes, members: list) -> dict:
    """
    解析宽表 Excel：
    | 日期 | Brant | Angel | ... |

    返回:
      {
        rows: [{date, member_name, member_id, content}],
        unbound: [{date, member_name, content}],
        errors: [str],
      }
    """
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    rows_iter = list(ws.iter_rows(values_only=True))
    if not rows_iter:
        raise ValueError("Excel 为空")

    header = [_cell_str(c) for c in rows_iter[0]]
    if not header or ("日期" not in header[0] and header[0].lower() not in ("date", "日期")):
        # 允许第一列叫 日期/Date
        if not header or header[0] not in ("日期", "Date", "date", "DATE"):
            raise ValueError("Excel 首列必须为「日期」")

    name_to_id = {}
    for m in members:
        name_to_id[m["name"].strip()] = m["id"]
        name_to_id[m["id"].strip()] = m["id"]

    member_cols = []  # (col_idx, header_name)
    for idx, name in enumerate(header[1:], start=1):
        if name:
            member_cols.append((idx, name))

    if not member_cols:
        raise ValueError("Excel 未找到成员列（第一行除日期外的列名）")

    parsed = []
    unbound = []
    errors = []
    seen_keys = set()

    for row_no, row in enumerate(rows_iter[1:], start=2):
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        report_date = _parse_date(row[0] if len(row) > 0 else None)
        if not report_date:
            errors.append(f"第{row_no}行：日期无效")
            continue

        for col_idx, member_name in member_cols:
            content = _cell_str(row[col_idx] if col_idx < len(row) else None)
            if not content:
                # 空内容跳过，记 error（按 Spec）
                continue

            biz_key = f"{report_date}_{member_name}"
            if biz_key in seen_keys:
                raise ValueError(f"发现重复日报：{report_date} {member_name}（请检查Excel）")
            seen_keys.add(biz_key)

            member_id = name_to_id.get(member_name)
            if not member_id:
                unbound.append({
                    "date": report_date,
                    "member_name": member_name,
                    "content": content,
                })
                continue

            parsed.append({
                "date": report_date,
                "member_name": member_name,
                "member_id": member_id,
                "content": content,
            })

    return {"rows": parsed, "unbound": unbound, "errors": errors}


# ========== Diff + 同步 ==========

def diff_and_sync(rows: list) -> dict:
    """
    对标准化行做 NEW / UPDATED / UNCHANGED
    返回结果统计 + 需 AI 分析的 report_id 列表
    """
    dates = sorted({r["date"] for r in rows})
    existing = get_daily_reports_by_dates(dates)

    new_items = []
    updated_items = []
    unchanged = 0
    analyze_ids = []

    for r in rows:
        key = f"{r['date']}_{r['member_id']}"
        h = content_hash(r["content"])
        old = existing.get(key)

        if not old:
            rid = insert_daily_report(r["date"], r["member_id"], r["content"], h, operator="excel_import")
            new_items.append({"report_id": rid, **r})
            analyze_ids.append(rid)
            continue

        if old.get("content_hash") == h:
            unchanged += 1
            continue

        update_daily_report(old["id"], r["content"], h, old.get("content") or "")
        updated_items.append({"report_id": old["id"], "old_content": old.get("content"), **r})
        analyze_ids.append(old["id"])

    return {
        "new": new_items,
        "updated": updated_items,
        "unchanged": unchanged,
        "analyze_ids": analyze_ids,
    }


def run_ai_analysis_for_reports(report_rows: list):
    """对 NEW/UPDATED 日报触发标签分析。report_rows: [{report_id, content, member_id, date}]"""
    for item in report_rows:
        try:
            analysis = analyze_daily_report(
                content=item["content"],
                member_name=item.get("member_name") or item.get("member_id"),
                report_date=item.get("date"),
            )
            upsert_daily_report_analysis(
                report_id=item["report_id"],
                skills=analysis.get("skills") or [],
                projects=analysis.get("projects") or [],
                activity_type=analysis.get("activity_type") or "开发",
                difficulty=int(analysis.get("difficulty") or 3),
                impact_score=float(analysis.get("impact_score") or 0),
                analysis_json=analysis,
                version=1,
            )
        except Exception as e:
            print(f"[daily_report] 分析失败 report_id={item.get('report_id')}: {e}")


# ========== 导入任务 ==========

def start_import_task(file_name: str, file_bytes: bytes) -> dict:
    task_id = create_daily_import_task(file_name)
    thread = threading.Thread(
        target=_run_import_task,
        args=(task_id, file_name, file_bytes),
        daemon=True,
    )
    thread.start()
    return {"task_id": str(task_id), "status": "processing"}


def _run_import_task(task_id: int, file_name: str, file_bytes: bytes):
    try:
        members = get_all_members()
        parsed = parse_daily_report_excel(file_bytes, members)
        rows = parsed["rows"]
        unbound = parsed["unbound"]
        parse_errors = parsed["errors"]

        # 空内容未计入 rows；对成员列为空的不强制 error_count
        empty_skipped = 0

        if not rows and not unbound:
            update_daily_import_task(
                task_id,
                status="failed",
                total_count=0,
                error_count=len(parse_errors) or 1,
                message="未解析到有效日报数据",
                result_json={"errors": parse_errors or ["未解析到有效日报数据"], "unbound": []},
            )
            return

        sync = diff_and_sync(rows)

        # AI 分析 NEW + UPDATED
        to_analyze = [
            {"report_id": x["report_id"], "content": x["content"],
             "member_id": x["member_id"], "member_name": x.get("member_name"), "date": x["date"]}
            for x in (sync["new"] + sync["updated"])
        ]
        if to_analyze:
            update_daily_import_task(task_id, message="正在进行 AI 标签分析...")
            run_ai_analysis_for_reports(to_analyze)

        result = {
            "new": len(sync["new"]),
            "updated": len(sync["updated"]),
            "unchanged": sync["unchanged"],
            "errors": parse_errors,
            "unbound": unbound,
            "mock_mode": is_mock_mode(),
            "analyzed": len(to_analyze),
        }
        update_daily_import_task(
            task_id,
            status="success",
            total_count=len(rows) + len(unbound),
            new_count=len(sync["new"]),
            update_count=len(sync["updated"]),
            skip_count=sync["unchanged"],
            error_count=len(parse_errors) + len(unbound),
            message="同步完成",
            result_json=result,
        )
    except Exception as e:
        update_daily_import_task(
            task_id,
            status="failed",
            message=str(e),
            error_count=1,
            result_json={"errors": [str(e)], "unbound": []},
        )


def get_import_task_status(task_id: int) -> Optional[dict]:
    task = get_daily_import_task(task_id)
    if not task:
        return None
    result = task.get("result_json") or {}
    return {
        "task_id": str(task["id"]),
        "file_name": task.get("file_name"),
        "status": task.get("status"),
        "message": task.get("message"),
        "total_count": task.get("total_count") or 0,
        "new": task.get("new_count") or result.get("new", 0),
        "updated": task.get("update_count") or result.get("updated", 0),
        "unchanged": task.get("skip_count") or result.get("unchanged", 0),
        "errors": task.get("error_count") or 0,
        "result": result,
        "created_at": task.get("created_at"),
        "updated_at": task.get("updated_at"),
    }


def build_ai_native_report_evidence(days=30) -> dict:
    """
    聚合日报证据供 AI Native 角色竞争使用：
    { member_id: { projects: Counter, skills: Counter, days: n, impact: sum } }
    """
    rows = get_member_recent_report_summary(days=days, limit=2000)
    evidence = {}
    for r in rows:
        mid = r["member_id"]
        bucket = evidence.setdefault(mid, {
            "projects": Counter(),
            "skills": Counter(),
            "days": 0,
            "impact": 0.0,
            "snippets": [],
        })
        bucket["days"] += 1
        bucket["impact"] += float(r.get("impact_score") or 0)
        for p in (r.get("projects") or []):
            bucket["projects"][p] += 1
        for s in (r.get("skills") or []):
            bucket["skills"][s] += 1
        if len(bucket["snippets"]) < 8:
            bucket["snippets"].append(f"[{r['report_date']}] {r.get('content','')[:80]}")

    # Counter -> dict
    out = {}
    for mid, b in evidence.items():
        out[mid] = {
            "projects": dict(b["projects"].most_common(10)),
            "skills": dict(b["skills"].most_common(10)),
            "days": b["days"],
            "impact": round(b["impact"], 1),
            "snippets": b["snippets"],
        }
    return out


def list_rewrite_styles():
    return list_daily_report_style_prompts()


def save_rewrite_style(style_id, prompt, label=None):
    if not (prompt or "").strip():
        raise ValueError("提示词不能为空")
    row = update_daily_report_style_prompt(style_id, prompt, label=label)
    if not row:
        raise KeyError("风格不存在")
    return row


def rewrite_report_text(raw_text, style_id, prompt_override=None):
    text = (raw_text or "").strip()
    if not text:
        raise ValueError("请先输入工作描述")
    style = get_daily_report_style_prompt(style_id)
    if not style:
        raise KeyError("风格不存在")
    prompt = (prompt_override or "").strip() or style["prompt"]
    result = rewrite_daily_report(text, prompt, style_label=style.get("label"))
    return {
        "text": result,
        "style_id": style["id"],
        "style_label": style.get("label"),
        "degraded": last_call_degraded() or is_mock_mode(),
    }


def ingest_generated_report(report_date, member_id, content):
    """一键转入：当天该成员已有日报则追加，否则新增。"""
    day = parse_day(report_date)
    if not day:
        raise ValueError("请选择有效日期")
    text = (content or "").strip()
    if not text:
        raise ValueError("日报内容不能为空")
    member = get_member(member_id)
    if not member:
        raise KeyError("成员不存在")

    existing = get_daily_report_by_key(day, member_id)
    if existing:
        old = (existing.get("content") or "").rstrip()
        merged = f"{old}\n\n{text}" if old else text
        if merged == (existing.get("content") or ""):
            return {
                "report_id": existing["id"],
                "action": "unchanged",
                "version": existing.get("version") or 1,
                "content": existing.get("content") or "",
                "report_date": day,
                "member_id": member_id,
                "member_name": member.get("name") or member_id,
            }
        h = content_hash(merged)
        version = update_daily_report(
            existing["id"], merged, h, existing.get("content") or "",
            operator="rewrite_ingest",
        )
        report_id = existing["id"]
        action = "appended"
        final_content = merged
    else:
        h = content_hash(text)
        report_id = insert_daily_report(
            day, member_id, text, h, operator="rewrite_ingest",
        )
        version = 1
        action = "created"
        final_content = text

    run_ai_analysis_for_reports([{
        "report_id": report_id,
        "content": final_content,
        "member_id": member_id,
        "member_name": member.get("name"),
        "date": day,
    }])
    return {
        "report_id": report_id,
        "action": action,
        "version": version,
        "content": final_content,
        "report_date": day,
        "member_id": member_id,
        "member_name": member.get("name") or member_id,
    }
